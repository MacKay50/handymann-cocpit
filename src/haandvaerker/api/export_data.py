from __future__ import annotations
from datetime import date, datetime
from io import BytesIO
from typing import Optional
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ..dependencies import CompanyContextDep
from ..models.admin_deadline import AdminDeadline
from ..models.appointment import Appointment
from ..models.customer import Customer
from ..models.employee import Employee
from ..models.historical_offer import HistoricalOffer
from ..models.inbox_message import InboxMessage
from ..models.invoice import Invoice
from ..models.project import Project
from ..models.quote import Quote

router = APIRouter(prefix="/export", tags=["export"])

# ── Workbook styling ─────────────────────────────────────────────────────────
_HDR_FILL = PatternFill("solid", fgColor="1A1A2E")
_HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
_ALT_FILL = PatternFill("solid", fgColor="F7FAFC")
_THIN = Side(style="thin", color="E2E8F0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _workbook(headers: list[str], rows: list[list], sheet_name: str) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.freeze_panes = "A2"
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = _BORDER
    for i, row in enumerate(rows, start=2):
        ws.append(row)
        if i % 2 == 0:
            for cell in ws[i]:
                cell.fill = _ALT_FILL
        for cell in ws[i]:
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="center")
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 3, 8), 52)
    ws.row_dimensions[1].height = 20
    return wb


def _stream(wb: openpyxl.Workbook, filename: str) -> StreamingResponse:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Value helpers ─────────────────────────────────────────────────────────────
def _d(val: object) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        return date.fromisoformat(str(val)[:10])
    except ValueError:
        return None


def _f(val: object) -> Optional[float]:
    return round(float(val), 2) if val is not None else None


def _match(fields: list[str], q: Optional[str]) -> bool:
    if not q:
        return True
    ql = q.lower()
    return any(ql in (f or "").lower() for f in fields)


# ── Per-type handlers ─────────────────────────────────────────────────────────

def _invoices(
    session: Session, company_id: str, status: Optional[str], q: Optional[str]
) -> tuple[list[str], list[list]]:
    cust_map = {c.id: c.name for c in session.exec(
        select(Customer).where(Customer.company_id == company_id, Customer.active == True)  # noqa: E712
    ).all()}
    proj_map = {p.id: p.title for p in session.exec(
        select(Project).where(Project.company_id == company_id, Project.active == True)  # noqa: E712
    ).all()}
    stmt = select(Invoice).where(Invoice.company_id == company_id, Invoice.active == True)  # noqa: E712
    if status:
        stmt = stmt.where(Invoice.status == status)
    stmt = stmt.order_by(Invoice.invoice_number)
    headers = ["Fakturanr.", "Titel", "Kunde", "Projekt",
               "Udstedt", "Forfald", "Status", "Subtotal (DKK)", "Moms (DKK)", "Total (DKK)"]
    rows = []
    for inv in session.exec(stmt).all():
        cust = cust_map.get(inv.customer_id, "")
        proj = proj_map.get(inv.project_id, "")
        if not _match([inv.invoice_number, inv.title or "", cust, proj], q):
            continue
        rows.append([inv.invoice_number, inv.title or "", cust, proj,
                     _d(inv.issue_date), _d(inv.due_date), inv.status.value,
                     _f(inv.subtotal), _f(inv.vat_amount), _f(inv.total)])
    return headers, rows


def _quotes(
    session: Session, company_id: str, status: Optional[str], q: Optional[str]
) -> tuple[list[str], list[list]]:
    proj_rows = session.exec(
        select(Project).where(Project.company_id == company_id, Project.active == True)  # noqa: E712
    ).all()
    proj_map = {p.id: (p.title, p.customer_id) for p in proj_rows}
    cust_map = {c.id: c.name for c in session.exec(
        select(Customer).where(Customer.company_id == company_id, Customer.active == True)  # noqa: E712
    ).all()}
    stmt = select(Quote).where(Quote.company_id == company_id, Quote.active == True)  # noqa: E712
    if status:
        stmt = stmt.where(Quote.status == status)
    stmt = stmt.order_by(Quote.quote_number)
    headers = ["Tilbudsnr.", "Titel", "Projekt", "Kunde", "Status",
               "Gyldig til", "Subtotal (DKK)", "Moms (DKK)", "Total (DKK)", "Accepteret dato"]
    rows = []
    for qu in session.exec(stmt).all():
        ptitle, cid = proj_map.get(qu.project_id, ("", ""))
        cust = cust_map.get(cid, "")
        if not _match([qu.quote_number, qu.title or "", ptitle, cust], q):
            continue
        rows.append([qu.quote_number, qu.title or "", ptitle, cust, qu.status.value,
                     _d(qu.valid_until), _f(qu.subtotal), _f(qu.vat_amount), _f(qu.total),
                     _d(qu.accepted_at)])
    return headers, rows


def _projects(
    session: Session, company_id: str, status: Optional[str], q: Optional[str]
) -> tuple[list[str], list[list]]:
    cust_map = {c.id: c.name for c in session.exec(
        select(Customer).where(Customer.company_id == company_id, Customer.active == True)  # noqa: E712
    ).all()}
    stmt = select(Project).where(Project.company_id == company_id, Project.active == True)  # noqa: E712
    if status:
        stmt = stmt.where(Project.status == status)
    stmt = stmt.order_by(Project.title)
    headers = ["Titel", "Kunde", "Adresse", "Status", "Startdato", "Slutdato"]
    rows = []
    for proj in session.exec(stmt).all():
        cust = cust_map.get(proj.customer_id, "")
        if not _match([proj.title, cust, proj.address or ""], q):
            continue
        rows.append([proj.title, cust, proj.address or "", proj.status.value,
                     _d(proj.start_date), _d(proj.end_date)])
    return headers, rows


def _customers(
    session: Session, company_id: str, status: Optional[str], q: Optional[str]
) -> tuple[list[str], list[list]]:
    stmt = select(Customer).where(Customer.company_id == company_id, Customer.active == True)  # noqa: E712
    stmt = stmt.order_by(Customer.name)
    headers = ["Navn", "Email", "Tlf.", "Adresse"]
    rows = []
    for c in session.exec(stmt).all():
        if not _match([c.name, c.email or "", c.phone or "", c.address or ""], q):
            continue
        rows.append([c.name, c.email or "", c.phone or "", c.address or ""])
    return headers, rows


def _employees(
    session: Session, company_id: str, status: Optional[str], q: Optional[str]
) -> tuple[list[str], list[list]]:
    stmt = select(Employee).where(Employee.company_id == company_id, Employee.active == True)  # noqa: E712
    stmt = stmt.order_by(Employee.name)
    headers = ["Navn", "Rolle", "Timepris (DKK)", "Ansat dato"]
    rows = []
    for emp in session.exec(stmt).all():
        if not _match([emp.name, emp.role or ""], q):
            continue
        rows.append([emp.name, emp.role or "", _f(emp.default_hourly_rate), _d(emp.hired_date)])
    return headers, rows


def _inbox(
    session: Session, company_id: str, status: Optional[str], q: Optional[str]
) -> tuple[list[str], list[list]]:
    stmt = (select(InboxMessage)
            .where(InboxMessage.company_id == company_id, InboxMessage.active == True)  # noqa: E712
            .order_by(InboxMessage.received_at.desc()))
    if status:
        stmt = stmt.where(InboxMessage.status == status)
    headers = ["Afsender", "Email", "Tlf.", "Emne", "Kanal", "Modtaget", "Status"]
    rows = []
    for msg in session.exec(stmt).all():
        if not _match([msg.sender_name or "", msg.sender_email or "", msg.subject or ""], q):
            continue
        rows.append([msg.sender_name or "", msg.sender_email or "", msg.sender_phone or "",
                     msg.subject or "", msg.source.value,
                     _d(msg.received_at), msg.status.value])
    return headers, rows


def _deadlines(
    session: Session, company_id: str, status: Optional[str], q: Optional[str]
) -> tuple[list[str], list[list]]:
    stmt = (select(AdminDeadline)
            .where(AdminDeadline.company_id == company_id, AdminDeadline.active == True)  # noqa: E712
            .order_by(AdminDeadline.due_date))
    if status:
        stmt = stmt.where(AdminDeadline.status == status)
    headers = ["Titel", "Kategori", "Forfaldsdato", "Status", "Noter"]
    rows = []
    for dl in session.exec(stmt).all():
        if not _match([dl.title, dl.notes or ""], q):
            continue
        rows.append([dl.title, dl.category.value, _d(dl.due_date), dl.status.value, dl.notes or ""])
    return headers, rows


def _appointments(
    session: Session, company_id: str, status: Optional[str], q: Optional[str]
) -> tuple[list[str], list[list]]:
    cust_map = {c.id: c.name for c in session.exec(
        select(Customer).where(Customer.company_id == company_id, Customer.active == True)  # noqa: E712
    ).all()}
    emp_map = {e.id: e.name for e in session.exec(
        select(Employee).where(Employee.company_id == company_id, Employee.active == True)  # noqa: E712
    ).all()}
    stmt = (select(Appointment)
            .where(Appointment.company_id == company_id, Appointment.active == True)  # noqa: E712
            .order_by(Appointment.start_datetime))
    if status:
        stmt = stmt.where(Appointment.status == status)
    headers = ["Titel", "Type", "Start", "Slut", "Sted", "Kunde", "Medarbejder", "Status"]
    rows = []
    for appt in session.exec(stmt).all():
        cust = cust_map.get(appt.customer_id or "", "")
        emp  = emp_map.get(appt.employee_id or "", "")
        if not _match([appt.title, appt.location or "", cust], q):
            continue
        rows.append([appt.title, appt.appointment_type.value,
                     _d(appt.start_datetime), _d(appt.end_datetime),
                     appt.location or "", cust, emp, appt.status.value])
    return headers, rows


def _erfaringsbank(
    session: Session, company_id: str, status: Optional[str], q: Optional[str]
) -> tuple[list[str], list[list]]:
    stmt = select(HistoricalOffer).where(HistoricalOffer.active == True)  # noqa: E712
    if status:
        stmt = stmt.where(HistoricalOffer.extraction_status == status)
    stmt = stmt.order_by(HistoricalOffer.year.desc())
    headers = ["Titel", "Jobbtype", "Areal m2", "Pris ekskl. moms (DKK)",
               "Pris/m2 (DKK)", "År", "Status"]
    rows = []
    for offer in session.exec(stmt).all():
        title = offer.title or offer.source_file_path.replace("\\", "/").split("/")[-1]
        ppm2 = round(offer.price_ex_vat / offer.area_m2, 2) \
               if offer.price_ex_vat and offer.area_m2 and offer.area_m2 > 0 else None
        if not _match([title, offer.job_type or "", offer.building_type or ""], q):
            continue
        rows.append([title, offer.job_type or "", _f(offer.area_m2),
                     _f(offer.price_ex_vat), ppm2, offer.year,
                     offer.extraction_status.value])
    return headers, rows


# ── Dispatch ──────────────────────────────────────────────────────────────────
_HANDLERS = {
    "invoices":      _invoices,
    "quotes":        _quotes,
    "projects":      _projects,
    "customers":     _customers,
    "employees":     _employees,
    "inbox":         _inbox,
    "deadlines":     _deadlines,
    "appointments":  _appointments,
    "erfaringsbank": _erfaringsbank,
}

_SHEET_NAMES = {
    "invoices": "Fakturaer", "quotes": "Tilbud", "projects": "Projekter",
    "customers": "Kunder", "employees": "Medarbejdere", "inbox": "Indbakke",
    "deadlines": "Frister", "appointments": "Aftaler", "erfaringsbank": "Erfaringsbank",
}


@router.get("/xlsx")
def export_xlsx(
    ctx: CompanyContextDep,
    type: str = Query(...),
    status: Optional[str] = Query(None),
    q: Optional[str] = Query(None),
) -> StreamingResponse:
    if type not in _HANDLERS:
        raise HTTPException(400, f"Ukendt type: {type}. Gyldige: {', '.join(_HANDLERS)}")
    headers, rows = _HANDLERS[type](ctx.session, ctx.company_id, status, q)
    wb = _workbook(headers, rows, _SHEET_NAMES.get(type, type.capitalize()))
    return _stream(wb, f"{type}_{date.today().isoformat()}.xlsx")
