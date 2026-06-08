from __future__ import annotations

import io
import pathlib
import pytest

from haandvaerker.services.danish_csv import (
    decode_csv_bytes,
    parse_danish_amount_ore,
    parse_danish_date,
    parse_danske_bank_csv,
    parse_economic_invoice_csv,
    parse_economic_invoice_xlsx,
)


@pytest.fixture
def fixtures_dir() -> pathlib.Path:
    return pathlib.Path(__file__).parent / "fixtures"


# --- parse_danish_amount_ore ---

def test_parse_danish_amount_ore_positive_thousands():
    assert parse_danish_amount_ore("1.234,56") == 123456


def test_parse_danish_amount_ore_negative():
    assert parse_danish_amount_ore("-1.234,56") == -123456


def test_parse_danish_amount_ore_small():
    assert parse_danish_amount_ore("0,01") == 1


def test_parse_danish_amount_ore_no_thousands_separator():
    assert parse_danish_amount_ore("12345,67") == 1234567


def test_parse_danish_amount_ore_large_amount():
    # Regression: float precision on amounts >= 100,000 DKK
    assert parse_danish_amount_ore("100.000,00") == 10000000


def test_parse_danish_amount_ore_invalid_raises():
    with pytest.raises(ValueError):
        parse_danish_amount_ore("UGYLDIG")


def test_parse_danish_amount_ore_whitespace_stripped():
    assert parse_danish_amount_ore("  1.234,56  ") == 123456


# --- parse_danish_date ---

def test_parse_danish_date():
    from datetime import date
    assert parse_danish_date("14-03-2026") == date(2026, 3, 14)


def test_parse_danish_date_first_of_month():
    from datetime import date
    assert parse_danish_date("01-01-2026") == date(2026, 1, 1)


def test_parse_danish_date_whitespace():
    from datetime import date
    assert parse_danish_date("  14-03-2026  ") == date(2026, 3, 14)


def test_parse_danish_date_invalid_raises():
    with pytest.raises(ValueError):
        parse_danish_date("2026-03-14")  # ISO format — wrong format


def test_parse_danish_date_garbage_raises():
    with pytest.raises(ValueError):
        parse_danish_date("not-a-date")


# --- decode_csv_bytes ---

def test_decode_csv_bytes_utf8():
    raw = "Dato;Tekst\n14-03-2026;test".encode("utf-8")
    result = decode_csv_bytes(raw)
    assert result.startswith("Dato;Tekst")


def test_decode_csv_bytes_cp1252():
    raw = "æøå".encode("cp1252")
    assert decode_csv_bytes(raw) == "æøå"


def test_decode_csv_bytes_cp1252_description():
    raw = "Overførsel fra Hansen Byggeri ApS".encode("cp1252")
    assert decode_csv_bytes(raw) == "Overførsel fra Hansen Byggeri ApS"


# --- parse_danske_bank_csv ---

def test_parse_danske_bank_csv_5_rows(fixtures_dir: pathlib.Path):
    content = (fixtures_dir / "danske_bank_sample.csv").read_text(encoding="utf-8")
    rows, errors = parse_danske_bank_csv(content, company_id="test-co")
    assert errors == []
    assert len(rows) == 5
    # 10.000,00 DKK = 1000000 øre
    assert rows[0].amount_ore == 1000000
    assert rows[0].description == "OVERF HANSEN BYGGERI"


def test_parse_danske_bank_csv_import_hash_present(fixtures_dir: pathlib.Path):
    content = (fixtures_dir / "danske_bank_sample.csv").read_text(encoding="utf-8")
    rows, errors = parse_danske_bank_csv(content, company_id="test-co")
    assert errors == []
    for row in rows:
        assert len(row.import_hash) == 64  # SHA-256 hex


def test_parse_danske_bank_csv_company_id_set(fixtures_dir: pathlib.Path):
    content = (fixtures_dir / "danske_bank_sample.csv").read_text(encoding="utf-8")
    rows, errors = parse_danske_bank_csv(content, company_id="my-company")
    assert errors == []
    for row in rows:
        assert row.company_id == "my-company"


def test_parse_danske_bank_csv_malformed_returns_errors(fixtures_dir: pathlib.Path):
    content = (fixtures_dir / "danske_bank_malformed.csv").read_text(encoding="utf-8")
    rows, errors = parse_danske_bank_csv(content, "test-co")
    assert len(errors) >= 1
    # Row 3 is the bad one (0-indexed row 2 = line 4 with header)
    assert any("4" in e or "Linje 4" in e for e in errors)


def test_parse_danske_bank_csv_malformed_no_rows_returned(fixtures_dir: pathlib.Path):
    # When there are errors, rows should be empty (all-or-nothing)
    content = (fixtures_dir / "danske_bank_malformed.csv").read_text(encoding="utf-8")
    rows, errors = parse_danske_bank_csv(content, "test-co")
    assert len(errors) >= 1
    assert rows == []


def test_parse_danske_bank_csv_with_valuta_column():
    # Danske Bank export with Valuta column after description (5-col variant)
    content = (
        "Bogfoeringsdato;Tekst;Beloeb;Valuta;Saldo\n"
        "28-05-2026;Betaling Hansen;5.000,00;DKK;15.000,00\n"
        "27-05-2026;El-regning;-1.234,56;DKK;10.000,00\n"
    )
    rows, errors = parse_danske_bank_csv(content, "test-co")
    assert errors == [], errors
    assert len(rows) == 2
    assert rows[0].amount_ore == 500000
    assert rows[0].description == "Betaling Hansen"
    assert rows[1].amount_ore == -123456


def test_parse_danske_bank_csv_valuta_before_amount():
    # Variant: Bogfoeringsdato;Tekst;Valuta;Beloeb;Saldo
    content = (
        "Bogfoeringsdato;Tekst;Valuta;Beloeb;Saldo\n"
        "28-05-2026;Overf ABC ApS;DKK;12.500,00;30.000,00\n"
    )
    rows, errors = parse_danske_bank_csv(content, "test-co")
    assert errors == [], errors
    assert len(rows) == 1
    assert rows[0].amount_ore == 1250000
    assert rows[0].description == "Overf ABC ApS"


def test_parse_danske_bank_csv_with_vaerdidato():
    # 6-col variant: Bogfoeringsdato;Vaerdidato;Tekst;Beloeb;Valuta;Saldo
    content = (
        "Bogfoeringsdato;Vaerdidato;Tekst;Beloeb;Valuta;Saldo\n"
        "28-05-2026;28-05-2026;Løn maj;25.000,00;DKK;50.000,00\n"
    )
    rows, errors = parse_danske_bank_csv(content, "test-co")
    assert errors == [], errors
    assert len(rows) == 1
    assert rows[0].amount_ore == 2500000
    assert rows[0].description == "Løn maj"


def test_parse_danske_bank_csv_dot_date_with_valuta():
    # Real-world format: DD.MM.YYYY dates combined with Valuta column
    content = (
        "Bogfoeringsdato;Tekst;Beloeb;Valuta;Saldo\n"
        "28.05.2026;Betaling;7.500,50;DKK;22.500,50\n"
    )
    rows, errors = parse_danske_bank_csv(content, "test-co")
    assert errors == [], errors
    assert len(rows) == 1
    assert rows[0].amount_ore == 750050
    from datetime import date
    assert rows[0].transaction_date == date(2026, 5, 28)


# --- parse_economic_invoice_csv ---

def test_parse_economic_invoice_csv_5_rows(fixtures_dir: pathlib.Path):
    content = (fixtures_dir / "economic_invoices_sample.csv").read_text(encoding="utf-8")
    rows, errors = parse_economic_invoice_csv(content, company_id="test-co")
    assert errors == []
    assert len(rows) == 5


def test_parse_economic_invoice_csv_first_row(fixtures_dir: pathlib.Path):
    from datetime import date
    content = (fixtures_dir / "economic_invoices_sample.csv").read_text(encoding="utf-8")
    rows, errors = parse_economic_invoice_csv(content, company_id="test-co")
    assert errors == []
    row = rows[0]
    assert row.economic_invoice_number == "10001"
    assert row.customer_name == "Hansen Byggeri"
    assert row.gross_amount_ore == 1000000  # 10.000,00 DKK = 1000000 øre
    assert row.due_date == date(2026, 3, 31)
    assert row.invoice_date == date(2026, 3, 1)


def test_parse_economic_invoice_csv_company_id_set(fixtures_dir: pathlib.Path):
    content = (fixtures_dir / "economic_invoices_sample.csv").read_text(encoding="utf-8")
    rows, errors = parse_economic_invoice_csv(content, company_id="my-firm")
    assert errors == []
    for row in rows:
        assert row.company_id == "my-firm"


# --- parse_economic_invoice_xlsx ---

def _make_invoice_xlsx(rows_data: list[tuple]) -> bytes:
    """Build a minimal in-memory xlsx with header + data rows."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Fakturanummer", "Debitor", "Netto", "Moms", "Brutto", "Forfaldsdato", "Bogfoeringsdato"])
    for row in rows_data:
        ws.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_economic_invoice_xlsx_string_amounts():
    from datetime import date
    raw = _make_invoice_xlsx([
        ("10001", "Hansen Byggeri", "8.000,00", "2.000,00", "10.000,00", "31-03-2026", "01-03-2026"),
    ])
    rows, errors = parse_economic_invoice_xlsx(raw, "test-co")
    assert errors == [], errors
    assert len(rows) == 1
    assert rows[0].economic_invoice_number == "10001"
    assert rows[0].customer_name == "Hansen Byggeri"
    assert rows[0].gross_amount_ore == 1000000
    assert rows[0].due_date == date(2026, 3, 31)


def test_parse_economic_invoice_xlsx_numeric_amounts():
    # e-conomic exports amounts as numbers in xlsx (not formatted strings)
    from datetime import date
    raw = _make_invoice_xlsx([
        ("10002", "Jensen Service", 4000.0, 1000.0, 5000.0, date(2026, 4, 30), date(2026, 4, 1)),
    ])
    rows, errors = parse_economic_invoice_xlsx(raw, "test-co")
    assert errors == [], errors
    assert rows[0].gross_amount_ore == 500000
    assert rows[0].due_date == date(2026, 4, 30)


def test_parse_economic_invoice_xlsx_skips_empty_rows():
    import openpyxl
    from datetime import date
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Fakturanummer", "Debitor", "Netto", "Moms", "Brutto", "Forfaldsdato", "Bogfoeringsdato"])
    ws.append(["10001", "Hansen", 800.0, 200.0, 1000.0, date(2026, 3, 31), date(2026, 3, 1)])
    ws.append([None, None, None, None, None, None, None])  # empty row
    ws.append(["10002", "Jensen", 800.0, 200.0, 1000.0, date(2026, 4, 30), date(2026, 4, 1)])
    buf = io.BytesIO()
    wb.save(buf)
    rows, errors = parse_economic_invoice_xlsx(buf.getvalue(), "test-co")
    assert errors == [], errors
    assert len(rows) == 2


def test_parse_economic_invoice_xlsx_header_detection():
    # xlsx with a company-name row before the column header (common in e-conomic exports)
    import openpyxl
    from datetime import date
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Min Virksomhed ApS", None, None, None, None, None, None])
    ws.append(["Fakturanummer", "Debitor", "Netto", "Moms", "Brutto", "Forfaldsdato", "Bogfoeringsdato"])
    ws.append(["10001", "Klima Service", 800.0, 200.0, 1000.0, date(2026, 3, 31), date(2026, 3, 1)])
    buf = io.BytesIO()
    wb.save(buf)
    rows, errors = parse_economic_invoice_xlsx(buf.getvalue(), "test-co")
    assert errors == [], errors
    assert len(rows) == 1
    assert rows[0].customer_name == "Klima Service"


def test_parse_economic_invoice_xlsx_dates_before_amounts():
    # e-conomic layout: Fakturanr | Dato | Forfaldsdato | Debitor | Netto | Moms | Brutto
    # Header names not in our detection sets → positional fallback puts dates in amount cols
    import openpyxl
    from datetime import date
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Fakturanr.", "Dato", "Forfaldsdato", "Debitornavn", "Netto", "Moms", "Brutto"])
    ws.append([10001, date(2026, 3, 1), date(2026, 3, 31), "Hansen Byggeri", 8000.0, 2000.0, 10000.0])
    ws.append([10002, date(2026, 4, 1), date(2026, 4, 30), "Jensen Service", 4000.0, 1000.0, 5000.0])
    buf = io.BytesIO()
    wb.save(buf)
    rows, errors = parse_economic_invoice_xlsx(buf.getvalue(), "test-co")
    assert errors == [], errors
    assert len(rows) == 2
    assert rows[0].gross_amount_ore == 1000000
    assert rows[0].due_date == date(2026, 3, 31)
    assert rows[0].invoice_date == date(2026, 3, 1)
    assert rows[0].customer_name == "Hansen Byggeri"


def test_parse_economic_invoice_xlsx_empty_returns_error():
    import openpyxl
    wb = openpyxl.Workbook()
    buf = io.BytesIO()
    wb.save(buf)
    rows, errors = parse_economic_invoice_xlsx(buf.getvalue(), "test-co")
    assert rows == []
    assert len(errors) >= 1
